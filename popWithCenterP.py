import arcpy.sa
import arcpy
from openpyxl import load_workbook
# change the below directory. this folder includes some scraps, shapefiles and population raster being used.
arcpy.env.workspace = r'C:\Users\MJ\Desktop\populationFunc\data'
arcpy.env.overwriteOutput = True
# change the directory to the path of excel file being used for input and output of data.
book = load_workbook(r'C:\Users\MJ\Desktop\populationFunc\data\pout.xlsx')
sheet = book.active

# function to create buffer around inputted xy
def creatbuf(po1, r):
    # below link stores the output of buffer. it is overwritten whenever runs.
    bufS = arcpy.Buffer_analysis(po1, "C:\\Users\\MJ\Desktop\\populationFunc\\data\\poshp", r)
    return (bufS)

# function to calculate population
def calpop(xy):
    r = [1 / 111, 1.5 / 111, 2 / 111, 2.5 / 111, 3 / 111, 4 / 111, 5 / 111, 7 / 111, 10 / 111]
    pop = []
    for i in range(0, 9):
        bufo = creatbuf(xy, r[i])
        popT = 'popT.dbf'
        rasterD = 'gurgaonPop.tif'
        popTable = arcpy.sa.ZonalStatisticsAsTable(bufo, "FID", rasterD, popT, "DATA", "SUM")
        # for printing a specific column from table
        fp = arcpy.SearchCursor(popTable)
        for f in fp:
            popVal = f.getValue("SUM")
        pop.append(popVal)
    return (pop)
        

latlng = []
i = 0
# to retrieve values stored in excel file
for row in range(3, sheet.max_row+1):
    for column in "BC":
        cell_name = "{}{}".format(column, row)
        latlng.append(sheet[cell_name].value)
    # GCS SRID
    spatial_reference = arcpy.SpatialReference(4326)
    # Point
    p1 = arcpy.Point(latlng[i+1], latlng[i])
    pnt_geometry = arcpy.PointGeometry(p1, spatial_reference)
    print(i,  "--", latlng[i+1], latlng[i])
    i = i+2
    popo = calpop(pnt_geometry)
    p = 0
    for column in "DEFGHIJKL":
        cell_name = "{}{}".format(column, row)
        sheet[cell_name].value = int(popo[p])
        print(int(popo[p]))
        p = p + 1

# change below directory to the output excel file...
book.save(r'C:\Users\MJ\Desktop\populationFunc\data\pout.xlsx')
        

