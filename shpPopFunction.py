import arcpy
from openpyxl import load_workbook
# workspace directory
arcpy.env.workspace = r'C:\Users\MJ\Desktop\populationFunc\data'
arcpy.env.overwriteOutput = True


# Area Calculation
def calarea(path):
    arcpy.AddGeometryAttributes_management(path, "AREA_GEODESIC", "KILOMETERS", "SQUARE_KILOMETERS")
    fi = arcpy.SearchCursor(path)
    for fa in fi:
        return fa.getValue("AREA_GEO")


# Population Calculation
def calpop(path):
    popT = 'popT.dbf'
    rasterD = 'gurgaonPop.tif'
    popTable = arcpy.sa.ZonalStatisticsAsTable(path, "FID", rasterD, popT, "DATA", "SUM")
    # for printing a specific column from table
    fp = arcpy.SearchCursor(popTable)
    for f in fp:
        popVal = f.getValue("SUM")
    return popVal


# Inputs
path = input("Enter Locality name with complete path and extension: ")
lName = input("Enter Locality Name: ")
areao = calarea(path)
popo = calpop(path)
print(popo)
print(areao)

# Saving in Excel
# path to open excel file where output is saved,
book = load_workbook(r'C:\Users\MJ\Desktop\populationFunc\data\det.xlsx')
sheet = book.active
vals = [lName, areao, popo]
i = 0
for row in range(sheet.max_row + 1, sheet.max_row + 2):
    for column in "ABC":
        cell_name = "{}{}".format(column, row)
        sheet[cell_name].value = vals[i]
        i = i + 1

# after writing the output values, excel is saved at the following path.
book.save(r'C:\Users\MJ\Desktop\populationFunc\data\det.xlsx')
