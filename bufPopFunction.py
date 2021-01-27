import arcpy.sa
import arcpy
from openpyxl import load_workbook
# below directory is workspace. Includes scrap and population raster and excel files for input and output of data,
arcpy.env.workspace = r'C:\Users\MJ\Desktop\populationFunc\data'
arcpy.env.overwriteOutput = True
# Link to the excel file for data input output
book = load_workbook(r'C:\Users\MJ\Desktop\populationFunc\data\det.xlsx')
sheet = book.active

# Population Calculation Function
def calpop(path):
    popT = 'popT.dbf'
    rasterD = 'gurgaonPop.tif'
    popTable = arcpy.sa.ZonalStatisticsAsTable(path, "FID", rasterD, popT, "DATA", "SUM")
    # for printing a specific column from table
    fp = arcpy.SearchCursor(popTable)
    for f in fp:
        popVal = f.getValue("SUM")
    return popVal


# Buffer around inputted Values
def creatbuf(po1, r):
    # workspace directory path... change it accordingly...popshp is the buffer output.
    bufS = arcpy.Buffer_analysis(po1, "C:\\Users\\MJ\Desktop\\populationFunc\\data\\poshp", r)
    return (bufS)


# Area Calculation
def calarea(bufS):
    arcpy.AddGeometryAttributes_management(bufS, "AREA_GEODESIC", "KILOMETERS", "SQUARE_KILOMETERS")
    fi = arcpy.SearchCursor(bufS)
    for fa in fi:
        area = fa.getValue("AREA_GEO")
    return (area)


# inputs
lName = input("Enter Localilty / area Name: ")
long = float(input("Enter longitude N value e.g: 77.xxxxx: "))
lat = float(input("Enter latitude E Value e.g: 28.xxxxx:  "))
radius = float(input("Enter Radius value in km: "))
# SRID and lat long are GCS (Decimal Degree) so dd = radius/111
radius = radius / 111
# GCS SRID
spatial_reference = arcpy.SpatialReference(4326)
# Point
p1 = arcpy.Point(28.481564, 76.973758)
pnt_geometry = arcpy.PointGeometry(p1, spatial_reference)
bufo = creatbuf(pnt_geometry, radius)
areao = calarea(bufo)
popo = calpop(bufo)


# print(pnt_geometry)
# print(areao)
# print(popo)


# Saving output in excel
vals = [lName, areao, popo]
i = 0
for row in range(sheet.max_row + 1, sheet.max_row + 2):
    for column in "ABC":
        cell_name = "{}{}".format(column, row)
        sheet[cell_name].value = vals[i]
        i = i + 1
# following excel file saves the output in written in excel
book.save(r'C:\Users\MJ\Desktop\populationFunc\data\det.xlsx')
